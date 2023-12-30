import os
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


class ParseExcel(object):

    def __init__(self, excel_path):
        self.font = None
        self.color_dict = None
        self.excel_path = excel_path
        self.workbook = load_workbook(self.excel_path, data_only=True)
        self.sheet = self.workbook.active

    def create_sheet_name(self, sheetname):
        self.workbook.create_sheet(sheetname)
        self.save_excel_file()

    def get_default_sheet(self):
        return self.sheet.title

    def set_sheet_by_index(self, sheet_index):
        sheet_name = self.workbook.sheetnames[sheet_index]
        self.sheet = self.workbook[sheet_name]
        return self.sheet

    def set_sheet_by_name(self, sheet_name):
        self.sheet = self.workbook[sheet_name]
        return self.sheet

    def get_max_row_no(self):
        return self.sheet.max_row

    def get_max_col_no(self):
        return self.sheet.max_column

    def get_min_row_no(self):
        return self.sheet.min_row

    def get_min_col_no(self):
        return self.sheet.min_column

    def get_all_rows(self):
        # return list(self.rows)
        return list(self.sheet.iter_rows())

    def get_all_cols(self):
        # return list(self.columns)
        return list(self.sheet.iter_cols())

    def get_single_row(self, row_no):
        return self.get_all_rows()[row_no]

    def get_single_col(self, col_no):
        return self.get_all_cols()[col_no]

    def get_cell(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no)

    def get_cell_content(self, row_no, col_no):
        return self.sheet.cell(row=row_no, column=col_no).value

    def write_cell_content(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content
        if font:
            self.font = Font(color=self.color_dict[font])
            self.sheet.cell(row=row_no, column=col_no).font = self.font
        self.workbook.save(self.excel_path)
        return self.sheet.cell(row=row_no, column=col_no).value

    def write_cell_content_colored(self, row_no, col_no, content, font=None):
        self.sheet.cell(row=row_no, column=col_no).value = content
        if font:
            fontx = PatternFill(patternType='solid', fgColor=font)
            self.sheet.cell(row=row_no, column=col_no).fill = fontx
        self.workbook.save(self.excel_path)
        return self.sheet.cell(row=row_no, column=col_no).value

    def write_cell_current_time(self, row_no, col_no):
        current_time = time.strftime('%Y-%m-%d %H:%M:%S')
        self.sheet.cell(row=row_no, column=col_no).value = str(current_time)
        self.workbook.save(self.excel_path)
        return self.sheet.cell(row=row_no, column=col_no).value

    def save_excel_file(self):
        self.workbook.save(self.excel_path)

    def get_valid_rows(self):
        print(self.sheet.cell)
        first_col = list(self.sheet.iter_cols())[0]  # 获取第一列
        for index, cell in enumerate(first_col):
            if index == len(first_col) - 1:
                return index
            elif cell.value is None:
                return index - 1

    def get_valid_cols(self):
        first_row = list(self.sheet.iter_rows())[0]  # 获取第一行
        for index, cell in enumerate(first_row):
            if index == len(first_row) - 1:
                return index + 1
            elif cell.value is None:
                return index

    def get_valid_data(self):
        row_count = self.get_valid_rows()
        col_count = self.get_valid_cols()
        if row_count and col_count:
            rows = self.get_all_rows()[1:row_count + 1]
            col_names = self.get_single_row(0)[:col_count]
            data_listxx = []
            for row in rows:
                row_dict = {}
                for i in range(0, col_count):
                    row_dict[col_names[i].value] = row[i].value
                data_listxx.append(row_dict)
            return data_listxx

    def get_all_rows_columns_data(self):
        col_count = self.get_valid_cols()
        rows = self.get_all_rows()[1:]
        col_names = self.get_single_row(0)[:col_count]
        data_listx = []
        for row in rows:
            row_dict = {}
            for i in range(0, col_count):
                row_dict[col_names[i].value] = row[i].value
            data_listx.append(row_dict)
        return data_listx

    def get_row_all_col_data(self, sheetname, columnname, pagename):
        self.set_sheet_by_name(sheetname)
        getall = self.get_all_rows_columns_data()
        # print(len(getall))
        # print(getall)
        xlem = []
        for xget in range(0, len(getall)):
            if getall[xget][columnname] == pagename:
                xlem.append(getall[xget])
        # print(xlem[2]['Elements'])
        return xlem

    def get_sheetnames(self):
        return self.workbook.sheetnames

    def rgbToInt(self, rgb):
        colorInt = rgb[0] + (rgb[1] * 256) + (rgb[2] * 256 * 256)
        return colorInt

    def sHyperlink(self, row_no, col_no, sshot, sshotpath):
        cell = self.sheet.cell(row=row_no, column=col_no)
        cell.value = sshot
        font = Font(underline="single", color="0563C1", name="Arial")
        cell.font = font
        cell.hyperlink = sshotpath
        self.workbook.save(self.excel_path)

    def getcomment(self, row_no, col_no, comments):
        column_letter = get_column_letter(col_no)
        cell_reference = f"{column_letter}{row_no}"
        comment_width = max(len(line) for line in comments.split("\n")) * 15
        comment = Comment(comments, "Author")
        self.sheet[cell_reference].comment = comment
        self.workbook.save(self.excel_path)