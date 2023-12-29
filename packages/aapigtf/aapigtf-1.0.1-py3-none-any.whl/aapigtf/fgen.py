import openpyxl
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.styles.borders import Side
from openpyxl.styles import Alignment


def create_excel_sheet_results(file_path, hTitle="ABC-CDG [ Portal ]",
                               sHtitle="Url: https://abc-cdg.com/api/v1/"):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "Web_Results"

    # Headers for the Excel sheet
    headers = ['Tc_no', 'Features', 'UserStory', 'Test_Case', 'Test_Steps', 'Test_data',
               'Actual_Results', 'Expected_Results', 'Status', 'Screenshots', 'IncidentIds']

    # Set header values in the first row

    # Set formatting for the title row (first row)
    title_font = Font(name='Tahoma', size=20, bold=True, color="FFFFFF", underline="single")
    title_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_border = Border(top=Side(border_style="thin", color="000000"),
                          left=Side(border_style="thin", color="000000"),
                          right=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the title row (first row)
    sheet.merge_cells('A1:K1')

    # Apply formatting for the title row (first row)
    title_cell = sheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    title_cell.border = title_border
    title_cell.value = hTitle

    # Set formatting for the subtitle row (second row)
    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    subtitle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    subtitle_border = Border(top=Side(border_style="thin", color="000000"),
                             left=Side(border_style="thin", color="000000"),
                             right=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the subtitle row (second row)
    sheet.merge_cells('E2:G2')

    # Apply formatting for the subtitle row (second row)
    subtitle_cell = sheet['E2']
    subtitle_cell.font = subtitle_font
    subtitle_cell.fill = subtitle_fill
    subtitle_cell.alignment = subtitle_alignment
    subtitle_cell.border = subtitle_border
    subtitle_cell.value = sHtitle

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    # Apply formatting for the subtitle row (second row)
    subtitle_cella = sheet['A2']
    subtitle_cella.font = subtitle_font
    subtitle_cella.fill = subtitle_fill
    subtitle_cella.alignment = subtitle_alignment
    subtitle_cella.border = subtitle_border
    subtitle_cella.value = 'Total Tests:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    subtitle_cellb = sheet['B2']
    subtitle_cellb.font = subtitle_font
    subtitle_cellb.fill = subtitle_fill
    subtitle_cellb.alignment = subtitle_alignment
    subtitle_cellb.border = subtitle_border
    subtitle_cellb.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_cellc = sheet['C2']
    subtitle_cellc.font = subtitle_font
    subtitle_cellc.fill = subtitle_fill
    subtitle_cellc.alignment = subtitle_alignment
    subtitle_cellc.border = subtitle_border
    subtitle_cellc.value = 'Total Passed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_celld = sheet['D2']
    subtitle_celld.font = subtitle_font
    subtitle_celld.fill = subtitle_fill
    subtitle_celld.alignment = subtitle_alignment
    subtitle_celld.border = subtitle_border
    subtitle_celld.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_cellh = sheet['H2']
    subtitle_cellh.font = subtitle_font
    subtitle_cellh.fill = subtitle_fill
    subtitle_cellh.alignment = subtitle_alignment
    subtitle_cellh.border = subtitle_border
    subtitle_cellh.value = 'Total Failed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_celli = sheet['I2']
    subtitle_celli.font = subtitle_font
    subtitle_celli.fill = subtitle_fill
    subtitle_celli.alignment = subtitle_alignment
    subtitle_celli.border = subtitle_border
    subtitle_celli.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellj = sheet['J2']
    subtitle_cellj.font = subtitle_font
    subtitle_cellj.fill = subtitle_fill
    subtitle_cellj.alignment = subtitle_alignment
    subtitle_cellj.border = subtitle_border
    subtitle_cellj.value = 'Total Skipped:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellk = sheet['k2']
    subtitle_cellk.font = subtitle_font
    subtitle_cellk.fill = subtitle_fill
    subtitle_cellk.alignment = subtitle_alignment
    subtitle_cellk.border = subtitle_border
    subtitle_cellk.value = '0'

    # Set formatting for the headers (third row)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

    sheet.append(headers)
    # Apply formatting for the headers (third row)
    for cell in sheet['A3:K3'][0]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)


def create_excel_sheet_for_object_repo(file_path, app=None):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "Object_Repo_BasePage"

    # Headers for the Excel sheet
    if app:
        headers = ['BasePage', 'Item', 'Locators', 'Elements1', 'Action', 'Description', 'FetchElements', 'Elements2']
    else:
        headers = ['BasePage', 'Item', 'Locators', 'Elements', 'Action', 'Description', 'FetchElements']

    # Set header values in the first row
    sheet.append(headers)

    # Apply formatting to the header row (underline, bold font, white text color, and red background)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)


def create_excel_sheet_for_testcase(file_path):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "FMS_Login"

    # Headers for the Excel sheet
    headers = ['Test_Case', 'TC_No', 'EndPoint', 'Method', 'Payload', 'Params', 'Test_Data',
               'Assertion_key', 'Format_Validation', 'Expected_API_Response', 'RunType', 'Case', 'Depend',
               'RequirementId', 'IncidentId']

    # Set header values in the first row
    sheet.append(headers)

    # Apply formatting to the header row (underline, bold font, white text color, and red background)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)


run_test = """import os
import cdxg
import urllib3
import argparse
from configparser import ConfigParser

urllib3.disable_warnings()
from pathlib import Path
from apixunit.common import generate_test_api

mypath = Path.cwd()
cObject = ConfigParser()
cObject.read(mypath / 'config.ini')
gen_file_path = []
test_case_data = cObject.get("test_data_xl", "test_case_data_xl")
if ',' in test_case_data:
    getallfiles = test_case_data.split(',')
    for xfilepath in getallfiles:
        gen_file_path.append(mypath / 'test_data' / xfilepath)
else:
    gen_file_path.append(mypath / 'test_data' / test_case_data)

if __name__ == '__main__':
    # run test dir
    # run test file
    parser = argparse.ArgumentParser(description='Add Arguments')
    parser.add_argument("--t", help="Add run type Tags")
    parser.add_argument("--s", help="Give Sheetname")
    parser.add_argument("--l", help="Give Executelines")
    args = parser.parse_args()
    tags_add = cObject['Tags']
    baseUrl = cObject['bURL']
    reporting = cObject['Reporting']
    tags_add['run_type'] = args.t
    with open('config.ini', 'w') as cf:
        cObject.write(cf)
    increment = 1
    for xfilepath in gen_file_path:
        if xfilepath.is_file():
            sheetname = args.s
            exeLine = args.l
            generate_test_api(xfilepath, sheetname, exeLine, reporting['title'], baseUrl['bUrl'], inc=increment)
            increment += 1

    cdxg.main(path="./test_dir/",
              base_url=baseUrl['bUrl'],
              title=reporting['title'],
              tester=reporting['tester'],
              report='result.html',
              open=True,
              debug=False,
              description=reporting['description'])
"""

conf_requires = """[Cookie_Headers]
xheaders = {"Content-Type": "application/x-www-form-urlencoded", "accept": "application/json", "X-User-ID": "xxxxxx"}

[App_Sys]
xapireport = apiReport

[test_data_xl]
test_case_data_xl = api_test_case_and_data.xlsx

[Tags]
run_type = Smoke

[bURL]
burl = {replace_api_url}


[Reporting]
title = {Give project title to reflect in html report}
tester = {tester_name}
description = {project description}
"""

apiconf_requires = """[apiEndpoint] 
nof = NetsOFF

[pdefinedget]
"""

test_requires = """aapigtf==1.0.1
"""

readme_requires = """# Accelerated API Generic tests Framework
* API Endpoints to be tested

## Prerequisites
```
Pycharm - Community Edition
python3.10 and above
```

**oneFMS Swagger API:**
```
* URL : {https://<api_url_address>/api>}
```

**Quick Start Installation**
```
1.	Get latest python3.10 and above with pip installed [**must]
2.	Go to terminal and install library : ‘pip or pip3 install virtualenv’
3.	To isolate virtual environment, Type Command: ‘python or python3 -m venv apitests' (**apitests is name of virtual environment, can give any name of choice)
4.	To activate virtual environment, 
    a.	Type Command: ‘source apitests/bin/activate’ (For Linux/Mac)
    b.	Type Command: ‘apitests>>Scripts' >> Type : 'activate’ (For Windows)
5.	To deactivate virtual environment, Type Command: ‘deactivate’
6.	To install dependencies, Type Command: ‘pip install apixunit -U’
```
**Start new project (follow the steps to create project, run tests and see results)**
```
1.	Go to root or the directory where to create project : apixunit <project_name> apix
    a. project_name : Any name as your choice (ex : 'fmsapi')
    b. apix : api or apix (refers to api testing)
2. Go to folder  ‘fmsapi>’ – All project files/folders parked here.
3. Go back to root folder ‘fmsapi’
4. To install all necessary packages, need to run the tests, Type: ‘pip or pip3 install -r requirements.txt’
5. Once all done, Good to go with add api testcases
    a. To add testcases and testdata use 'fmsapi>test_data>web>api_test_case_and_case.xlsx' (Any file name as your choice)
```
**Exisiting project (follow the steps to add api testcases frequently, run tests and see results)**
```
1. Always activate isolated virtual environment mode to run tests: source apitests/bin/activate
2. Make sure it must be in a isolated virtual environment as ‘(apitests)’
3. Go back to root folder ‘fmsapi’
4. Good to go with adding new tests and run tests, Repeat the same steps whenever needed
    a. Add "[Cookie_Headers]" for Authorization or Authentication to run tests if needed or required:
        a. For example :  xheaders = {"Content-Type": "application/x-www-form-urlencoded", "accept": "application/json", 
        `"X-User-ID": "xxxxxx" or , "Authorization": "Bearer eyJhbGciOiJSUzI" etc...}
    b. For more details about Authorization or Authentication cookies, headers refer: Swagger api documentation
```

##### Folders with files and Others:
```
1. test_data[testcases written with testdata as excel format], 
2. reports[results in form of Html and excel formats], 
3. test_dir[Api test(test_api_fms.py inherit with data driven using files from test_data and py files from utils]
4. utils [validating the request and response of api using the actual and expected results given in testcases]
5. requirements.txt[defining the list of libraries, which needs to be installed, after python installed]
6. run.py [Initial test execution file]
```

###### Project Structure
```
-----------
    ├── reports                             <- Reports and logs are parked here as excel and html format.
    ├── test_data                           <- Folder contains all api testcases with test data.
    │   ├── json_data                       <- Auto generation of json responses to correlate and use for dependent cases.
    │   ├── mixdata                         <- Auto generation of json mixed responses to correlate and use for dependent cases.
    │   ├── reqdata                         <- Auto generation of json requests to correlate and use for dependent cases.
    │   ├── resdata                         <- Auto generation of json responses to correlate and use for dependent cases.
    │   ├── api_test_data.xlsx              <- Use this format to add the testcases and test data for exeution
    ├── test_dir                            <- Main source code folder
    │   ├── test_*.py                       <- Inherit all functions to start execute the tests.
    ├── utils                               <- Included all base actions, commonhelpers,..
    │       ├── endpointapi.ini             <- Contains api endpoint headers to used for test
    ├── config.ini                          <- Contains key-value pair for properties and sections for test
    ├── README.md                           <- README for users using this project instructions.
    ├── requirements.txt                    <- Contains mandatory py libraries to be used for this project.
    ├── run.py                              <- Execute this to start the fms api tests.
```

**Tagging argument**:
--t=: [Smoke, Regression, Skip] --> Its Mandatory used in 'RunType' column in testcase excelsheet to define tagging.
--s=: [Excelsheet name incase of multiple sheets] 
--l=: [excelsheet row number(starts from row 2, 1 used for define headers)] )

**Run the test and see results**
```
1.	Type to run single specific tests in excelsheet: **‘fmsapi>python run.py --t=Smoke --s='FMS_Vehicle_Management', --l=2 or 5 or 10’
2.  Type to run multiple tests in specific excelsheet(10 tests in a row): **‘fmsapi>python run.py --t=Smoke --s='FMS_Vehicle_Management', --l=2,11’
3.  Type to run all tests specific to excelsheeet: **‘fmsapi>python run.py --t=Smoke --s='FMS_Vehicle_Management’
4.  Type to run all tests not specific to excelsheeets : **‘fmsapi>python run.py --t=Smoke’
2.	After Executions, Results displays under reports folder[*.html, *.xlsx]
```
"""

samp_test = '''
import json
import os
from cdxg import file_data
from apixunit.api import apicall
from cdxg.logging import log
from pathlib import Path
from configparser import ConfigParser
from apixunit.common import get_tags, Create_New_Report, get_results, getepoint, get_sheetnames_excel, excel_row_data
from apixunit.apiexecutor import apiSteps

project_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))
mypath = Path.cwd()
config = ConfigParser()
config.read(os.path.join(project_path, 'config.ini'))
test_case_data = config.get("test_data_xl", "test_case_data_xl")
generate_test_ap =  os.path.join(project_path, 'test_data', test_case_data)
xHeaders = json.loads(config.get("Cookie_Headers", "xHeaders"))
xreport = config.get("App_Sys", "xapireport")
reportpath = Create_New_Report(report=xreport, sTitle="NGP >> Payments", hTitle="https://api.sit.zig.live/v1.0/payment-services/")
test_tags = config.get("Tags", "run_type")
line = 2
end_line = None


class Api_Onefms_FMS_NetsOFF(apicall):

	def start(self):
		self.maxDiff = None

	@file_data(str(generate_test_ap), line=line, end_line=end_line, sheet="FMS_NetsOFF")
	def test_FMS_NetsOFF_api(self, testcase, tcdef, xendpoint, xmethod, payload, params, testdata,
							exresults, scvarib, ctype, results, rtype, case, depend, sprints, incidents):
		"""*"""
		log.info("***" + str(tcdef) + "_" + str(testcase) + "***")
		cfilex = config.read(os.path.join(project_path, 'utils', 'endpointapi.ini'))
		if '/' in xendpoint:
			xpoint = str(xendpoint).split('/')
			rolepoint = getepoint(xpoint)  # xpoint[len(xpoint)-1]
		else:
			rolepoint = xendpoint
		roles = config.get("apiEndpoint", rolepoint)
		if ',' in roles:
			roles, cxUrl = roles.split(',')
		if rtype != "skip" and rtype == get_tags(rtype) and rtype is not None:
			gtdata = get_sheetnames_excel(generate_test_ap, itemdata=depend, ustory=tcdef)
			for xlen in range(0, len(gtdata)):
				getallx = excel_row_data(gtdata[xlen][0], depend)
				if depend is not None:
					apiSteps().execute_steps(getallx, reportpath, xHeaders)
				else:
					apiSteps().execute_steps(getallx, reportpath, xHeaders)
		else:
			get_results(reportpath, roles, testcase, xendpoint, payload, params, exresults, 'Testcase Skipped',
							results='SKIPPED', fontx='F4D25A', httpcodes=000, elapsed_secs=0.00, comments=scvarib,
							sprints=sprints, incidentids=incidents, alljson='')
			if rtype == "skip":
				self.xSkip(testcase + ": Testcase Skipped, Due to not much information")
			else:
				self.skipTest(reason="Test execution based on Tags :" + str(test_tags) + ": Excludes the rest")
'''