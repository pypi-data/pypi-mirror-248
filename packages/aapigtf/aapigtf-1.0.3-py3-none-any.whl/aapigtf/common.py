import json
import time
import shutil
import string
import random
import openpyxl
import re
import configparser
from pathlib import Path
from aapigtf.parseexcel import ParseExcel
from aapigtf.apivalid import Validations
from aapigtf.ghelper import update_values_nested
from datetime import date
from cdxg import testdata
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.styles.borders import Side
from openpyxl.styles import Alignment

mypath = Path.cwd()  # .parent
config = configparser.ConfigParser()
data_file_path = mypath / 'config.ini'
config.read(mypath / 'config.ini')


def dumpData(apiselect, getalljson):
    with open(apiselect, 'w', encoding='utf-8') as jsonfile:
        json.dump(getalljson, jsonfile, ensure_ascii=False, indent=4)
    time.sleep(1)
    # log.info(getalljson)
    with open(apiselect) as data_json_file:
        data_loaded = json.load(data_json_file)
    # print(data_loaded)
    return data_loaded


def data_required(datafile):
    # print('*******' + str(datafile) + '*********')
    # testdata_file = str(datafile) + '.json'
    with open(datafile) as data_json_file:
        data_loaded = json.load(data_json_file)
    # print(data_loaded)
    return data_loaded


def get_domain_id(datafile):
    getdata, get_data = data_required(datafile), None
    getdomain = getdata['data']['userInfo']['domain']
    for lenx in range(0, len(getdomain)):
        if getdomain[lenx]['current']:
            get_data = getdomain[lenx]['id']
            break
    return get_data


def get_data_json_info(getalljson, datafile):
    with open('./test_data/json_data/' + str(datafile) + '_json_data.json', 'w', encoding='utf-8') as jsonfile:
        json.dump(getalljson, jsonfile, ensure_ascii=False, indent=4)
    time.sleep(1)


def getcURL(methd, url, data, param):
    if param != 'None' and methd == 'GET':
        url = url + '?' + param

    if data == 'None' or data == {}:
        data = ''

    command = "curl -X '" + str(methd) + "' '" + str(url) + "'"
    for key, value in json.loads(config.get("Cookie_Headers", "xHeaders")).items():
        command += " -H '" + str(key) + ":" + str(value) + "'"
    if methd != 'GET':
        command += " -d '" + str(data) + "'"
    return command


def get_results(xlcreate, roles, testcase, tendpoint, testdata, testparams, exResults, acResults, results, fontx,
                httpcodes, elapsed_secs, comments='', sprints='', incidentids='', alljson=''):
    reportpath = ParseExcel(excel_path=xlcreate)
    reportpath.set_sheet_by_name('API_Results')
    get_total_rows = reportpath.get_max_row_no()
    reportpath.write_cell_content_colored(2, 2, '=COUNTA(A4:A10000)')
    reportpath.write_cell_content_colored(2, 4, '=COUNTIF(I4:I10000, "✅")')
    reportpath.write_cell_content_colored(2, 12, '=COUNTIF(I4:I10000, "❌")')
    reportpath.write_cell_content_colored(2, 14, '=COUNTIF(I4:I10000, "⚠️")')
    try:
        reportpath.write_cell_content_colored(get_total_rows + 1, 1, get_total_rows - 2)
        reportpath.write_cell_content_colored(get_total_rows + 1, 2, roles)
        reportpath.write_cell_content_colored(get_total_rows + 1, 3, testcase)
        reportpath.write_cell_content_colored(get_total_rows + 1, 4, tendpoint)
        reportpath.write_cell_content_colored(get_total_rows + 1, 5, testdata)
        reportpath.write_cell_content_colored(get_total_rows + 1, 6, testparams)
        reportpath.write_cell_content_colored(get_total_rows + 1, 7, str(exResults))
        reportpath.write_cell_content_colored(get_total_rows + 1, 8, str(acResults))
        reportpath.getcomment(get_total_rows + 1, 8, str(alljson))
        if results == 'PASSED':
            reportpath.write_cell_content_colored(get_total_rows + 1, 9, '✅', font=fontx)
        elif results == 'FAILED':
            reportpath.write_cell_content_colored(get_total_rows + 1, 9, '❌', font=fontx)
        else:
            reportpath.write_cell_content_colored(get_total_rows + 1, 9, '⚠️', font=fontx)
        # reportpath.write_cell_content_colored(get_total_rows + 1, 9, results, font=fontx)
        reportpath.write_cell_content_colored(get_total_rows + 1, 10, comments)
        reportpath.write_cell_content_colored(get_total_rows + 1, 11, httpcodes)
        reportpath.write_cell_content_colored(get_total_rows + 1, 12, elapsed_secs)
        reportpath.write_cell_content_colored(get_total_rows + 1, 13, sprints)
        reportpath.write_cell_content_colored(get_total_rows + 1, 14, incidentids)
    except Exception as e:
        if ValueError:
            for xna in acResults:
                reportpath.write_cell_content_colored(get_total_rows + 1, 5, xna)
        # print(str(e))


def Create_New_Report(report, sTitle, hTitle):
    otime = time.strftime("%Y-%m-%d_%H%M%S", time.localtime())
    reportfolderpath = report + "_" + otime + ".xlsx"
    # shutil.copy(mypath / 'reports' / 'api_results.xlsx', mypath / 'reports' / reportfolderpath)
    reportPath = mypath / 'reports' / reportfolderpath
    create_excel_sheet_results(file_path=reportPath, hTitle=sTitle, sHtitle=hTitle)
    return mypath / 'reports' / reportfolderpath


def get_tags(rtype='Smoke'):
    config.read(data_file_path)
    gettag = config['Tags']
    test_tags = gettag['run_type']
    if rtype is not None:
        rty = rtype.split(',')
        ttags = test_tags.split(',')
        if len(rty) == 1 and len(ttags) >= 1:
            if test_tags.find(rty[0]) != -1:
                t_xgs = None
                for txgs in ttags:
                    if txgs == rty[0]:
                        t_xgs = txgs
                return t_xgs
        else:
            xtype, txn = [], None
            for tgs in ttags:
                for rxt in rty:
                    if rxt == tgs:
                        txn = rxt
                        break
                if txn == tgs:
                    xtype.append('Y')
                else:
                    xtype.append('N')
            if 'Y' in list(set(xtype)):
                return rtype


def getTypesplit(xType, get_data, gtd, tdata):
    global charsx, gxdatax
    today = date.today()
    date_string = today.strftime('%Y_%m_%d')
    if xType.startswith('@'):
        gSplit = xType.split('^')
        for xlen in range(1, len(gSplit[1])):
            xSplit = gSplit[1].split(',')
            pSplit = gSplit[2].split(',')
            for xlenn in range(0, len(xSplit)):
                if xSplit[xlenn] == 'Addx':
                    gcspt = pSplit[xlenn].split('=')
                    if '|' in gcspt[1]:
                        getcharsx = gcspt[1].split('*')
                        mxchars = mixedcharacters(k=int(getcharsx[1]))
                    else:
                        mxchars = gcspt[1] + '_' + date_string
                    get_data[str(gcspt[0])] = str(mxchars)
                elif xSplit[xlenn] == 'Add':
                    gkey, gvalue = pSplit[xlenn].split('=')
                    if '|' in gvalue and '*' in gvalue and '.' not in gkey:
                        getcharsx = gvalue.split('*')
                        charsx = mixedcharacters(k=int(getcharsx[1]))
                        get_data[str(gkey)] = charsx
                    elif '.' in gkey:
                        cname, kname = gkey.split('.')
                        charsx = update_values_nested(get_data, cname, kname, gvalue)
                        get_data = charsx
                    else:
                        mxchars = gvalue
                        if mxchars == '[]':
                            charsx = []
                        else:
                            charsx = str(mxchars)
                        get_data[str(gkey)] = charsx
                elif xSplit[xlenn] == 'Extract':
                    gcspt = pSplit[xlenn].split('=')
                    # print(gcspt)
                    gxdatax = getExtract(xdata=gcspt[1], tdata=tdata)
                    get_data[str(gcspt[0])] = gxdatax
                elif xSplit[xlenn] == 'Dadd':
                    keyx, valuesx = pSplit[xlenn].split('=')
                    rpl = valuesx.replace('*', ',')
                    xType_dict = json.loads(rpl)
                    get_data[keyx] = xType_dict
                elif xSplit[xlenn] == 'Ladd':
                    grt = []
                    gcspt = pSplit[xlenn].split('=')
                    gmt = gcspt[1].split('-')
                    try:
                        for xgmt in gmt:
                            if '.' in xgmt:
                                grt.append(float(xgmt))
                            else:
                                grt.append(int(xgmt))
                    except Exception:
                        for xgmt in gmt:
                            grt.append(xgmt)
                    get_data[str(gcspt[0])] = grt
                else:
                    if xSplit[xlenn] == 'Remove':
                        if '=' in pSplit[xlenn]:
                            gcspt = pSplit[xlenn].split('=')
                            if '*' in gcspt[0]:
                                attrN = gcspt[0].split('*')
                                for xattr in attrN:
                                    gxdatax = getremove(xdata=gcspt[1], tdata=tdata, getn=get_data)
                                delgtx = attrN
                            else:
                                gxdatax = getremove(xdata=gcspt[1], tdata=tdata, getn=get_data)
                                delgtx = gcspt[0]
                            ggx = []
                            for xgtx in gxdatax:
                                if type(delgtx) == list:
                                    for delg in delgtx:
                                        del xgtx[delg]
                                        ggx.append(xgtx)
                                else:
                                    if type(xgtx) == dict:
                                        del xgtx[delgtx]
                                        ggx.append(xgtx)
                                    else:
                                        for xgt in xgtx:
                                            del xgt[delgtx]
                                            ggx.append(xgt)
                            if len(ggx) >= tdata:
                                ggx = random.sample(ggx, tdata)
                                # print(ggx)
                            if '|' in gcspt[1]:
                                aName = gcspt[1].split('|')
                                get_data[str(aName[3])] = ggx
                            else:
                                get_data[str(gcspt[1])] = ggx
                        else:
                            del get_data[pSplit[xlenn]]
            return json.dumps(dict(get_data))
    else:
        if gtd:
            return json.dumps(dict(get_data))
        else:
            return xType


def find_row_number_within_range(test_case_data, start_variable, end_variable):
    workbook = ParseExcel(test_case_data).workbook
    found_start = False
    found_end = False
    start_row = None
    end_row = None
    sheet_name_start = None
    sheet_name_end = None

    if end_variable is None or end_variable == "":
        end_variable = start_variable

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # print(sheet)
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # print(row_num, row)
            if row[1] == start_variable:
                found_start = True
                start_row = row_num
                sheet_name_start = sheet.title
            if row[1] == end_variable:
                found_end = True
                end_row = row_num
                sheet_name_end = sheet.title

            if found_start and found_end:
                break
        # print(start_row, end_row, sheet_name_start, sheet_name_end)
    workbook.close()
    if start_row is not None and end_row is not None and start_row <= end_row:
        return {'sLine': start_row, 'eLine': end_row, 'sheet_name': sheet_name_start}
    else:
        return "Start or end variable not found in the specified range.", 0
    # return start_row, end_row, sheet_name_start, sheet_name_end


def is_valid_value(value):
    if ',' in value:
        # Check if it's a comma-separated list of values
        values = value.split(',')
        for val in values:
            if not val.strip().isdigit():
                return False
        return True
    return value.isdigit()


def exeLiner(test_case_data, exeLine):
    global sName
    if is_valid_value(exeLine):
        if exeLine is not None:
            if ',' in exeLine:
                values = exeLine.split(',')
                gvalues = list(map(int, values))
                # print(f"The argument --l is a list of integers: {list(map(int, values))}")
                return ','.join(str(item) for item in gvalues)
        else:
            # print(f"The argument --l is an integer: {int(exeLine)}")
            return str(exeLine[0]) + ',' + str(exeLine[0])
    else:
        if ',' in exeLine:
            start_row, end_row = exeLine.split(',')
        else:
            start_row, end_row = exeLine, None
        getLine = find_row_number_within_range(test_case_data, start_row, end_row)
        exeLine = str(getLine['sLine']) + ',' + str(getLine['eLine'])
        sName = str(getLine['sheet_name'])
    return exeLine, sName


def generate_test_api(generate_test_api, sheetname=None, exeLine=None, sTitle=None, hTitle=None, inc=0):
    sLine, eLine = 2, None
    getxcl = ParseExcel(generate_test_api)
    getallxm = getxcl.get_sheetnames()
    trig = 'NO'
    if sheetname is None:
        trig = 'YES'
    elif sheetname is not None and ',' not in sheetname:
        if sheetname in getallxm:  # Check if the sheet name is in the getallxm list
            trig = 'YES'
    elif sheetname is not None and ',' in sheetname:
        sheetname_list = sheetname.split(',')  # Split the string into a list of sheet names
        for xsheetlist in sheetname_list:
            if xsheetlist in getallxm:  # Check if the sheet name is in the getallxm list
                trig = 'YES'
    else: pass
    # print(trig)
    if trig == 'YES':
        getxmall = []
        try:
            if sheetname:
                for y in eval(sheetname):
                    for x in range(0, len(getallxm)):
                        if x == y or getallxm[x] == y:
                            getxmall.append(getallxm[x])
                            break
                # Update getallxm with the filtered values
                getallxm = getxmall
        except Exception:
            sheetname_list = sheetname.split(',')  # Split the string into a list of sheet names
            for y in sheetname_list:
                if y in getallxm:  # Check if the sheet name is in the getallxm list
                    getxmall.append(y)
            # Update getallxm with the filtered values
            getallxm = getxmall

        if exeLine is not None and sheetname is None:
            exeLine = exeLiner(generate_test_api, exeLine)
            sLine, eLine = str(exeLine[0]).split(',')
            getallxm = [str(exeLine[1])]
        elif sheetname is not None and exeLine is None or sheetname is None and exeLine is None:
            sLine, eLine = sLine, eLine
        else:
            if ',' in exeLine:
                sLine, eLine = str(exeLine).split(',')
            else:
                sLine, eLine = exeLine, None
        pathdata_file = mypath / 'test_dir' / f'test_api{inc}.py'
        print(pathdata_file)
        with open(pathdata_file, 'w') as f:
            f.write('import json\n')
            f.write('import os\n')
            f.write('from cdxg import file_data\n')
            f.write('from aapigtf.api import apicall\n')
            f.write('from cdxg.logging import log\n')
            f.write('from pathlib import Path\n')
            f.write('from configparser import ConfigParser\n')
            f.write('from aapigtf.common import get_tags, Create_New_Report, get_results, getepoint, get_sheetnames_excel, excel_row_data\n')
            f.write('from aapigtf.apiexecutor import apiSteps\n\n')
            f.write('project_path = os.path.dirname(os.path.dirname(os.path.realpath(__file__)))\n')
            f.write('mypath = Path.cwd()\n')
            f.write('config = ConfigParser()\n')
            f.write("config.read(os.path.join(project_path, 'config.ini'))\n")
            f.write('test_case_data = config.get("test_data_xl", "test_case_data_xl")\n')
            f.write("generate_test_ap =  os.path.join(project_path, 'test_data', test_case_data)\n")
            f.write('xHeaders = json.loads(config.get("Cookie_Headers", "xHeaders"))\n')
            f.write('xreport = config.get("App_Sys", "xapireport")\n')
            f.write('reportpath = Create_New_Report(report=xreport, sTitle="' + sTitle + '", hTitle="' + hTitle + '")\n')
            f.write('test_tags = config.get("Tags", "run_type")\n')
            # f.write('base_url = config.get("bURL", "burl")\n')
            f.write('line = ' + str(sLine) + '\n')
            f.write('end_line = ' + str(eLine) + '\n\n\n')

            for getxl in getallxm:
                if getxl != 'Usage':
                    f.write('class Api_Onefms_' + str(getxl) + '(apicall):\n\n')
                    f.write('\tdef start(self):\n')
                    f.write('\t\tself.maxDiff = None\n\n')
                    f.write(
                        '\t@file_data(str("'+str(generate_test_api)+'"), line=line, end_line=end_line, sheet="' + str(getxl) + '")\n')
                    f.write(
                        '\tdef test_' + getxl + '_api(self, testcase, tcdef, xendpoint, xmethod, payload, params, testdata,\n')
                    f.write('\t\t\t\t\t\t\tscvarib, ctype, exresults, rtype, case, depend, sprints, incidents):\n')
                    f.write('\t\t"""*"""\n')
                    f.write('\t\tif rtype and case is not None:\n')
                    f.write('\t\t\txtype = rtype + "," + case\n')
                    f.write('\t\telse:\n')
                    f.write('\t\t\txtype = rtype if case is None else case\n')
                    f.write('\t\tif get_tags(xtype):\n')
                    f.write('\t\t\tlog.info("***[" + str(tcdef) + "]-[" + str(testcase) + "]-[" + str(get_tags(xtype)) + "]***")\n')
                    # f.write('\t\t\tlog.info("Test_Tags : " + str(get_tags(xtype)))\n')
                    f.write("\t\tconfig.read(os.path.join(project_path, 'utils', 'endpointapi.ini'))\n")
                    f.write("\t\tif '/' in xendpoint:\n")
                    f.write("\t\t\txpoint = str(xendpoint).split('/')\n")
                    f.write("\t\t\trolepoint = getepoint(xpoint)  # xpoint[len(xpoint)-1]\n")
                    f.write('\t\telse:\n')
                    f.write("\t\t\trolepoint = xendpoint\n")
                    f.write('\t\troles = config.get("apiEndpoint", rolepoint)\n')
                    f.write("\t\tif ',' in roles:\n")
                    f.write("\t\t\troles, cxUrl = roles.split(',')\n")
                    # f.write('\t\tif rtype != "skip" and rtype == get_tags(rtype) and rtype is not None:\n')
                    f.write('\t\tif "skip" not in xtype and xtype == get_tags(xtype) and xtype is not None:\n')
                    f.write("\t\t\tgtdata = get_sheetnames_excel('"+str(generate_test_api)+"', itemdata=depend, ustory=tcdef)\n")
                    f.write("\t\t\tfor xlen in range(0, len(gtdata)):\n")
                    f.write("\t\t\t\tgetallx = excel_row_data(gtdata[xlen][0], depend)\n")
                    f.write("\t\t\t\tif depend is not None:\n")
                    f.write("\t\t\t\t\tapiSteps().execute_steps(getallx, reportpath, xHeaders)\n")
                    f.write("\t\t\t\telse:\n")
                    f.write("\t\t\t\t\tapiSteps().execute_steps(getallx, reportpath, xHeaders)\n")
                    f.write('\t\telse:\n')
                    f.write(
                        "\t\t\tget_results(reportpath, roles, testcase, xendpoint, payload, params, exresults, 'Testcase Skipped',\n")
                    f.write(
                        "\t\t\t\t\t\t\tresults='SKIPPED', fontx='F4D25A', httpcodes=000, elapsed_secs=0.00, comments=scvarib,\n")
                    f.write("\t\t\t\t\t\t\tsprints=sprints, incidentids=incidents, alljson='')\n")
                    f.write('\t\t\tif rtype == "skip":\n')
                    f.write('\t\t\t\tself.xSkip(testcase + ": Testcase Skipped, Due to not much information")\n')
                    f.write('\t\t\telse:\n')
                    f.write(
                        '\t\t\t\tself.skipTest(reason="Test execution based on Tags :" + str(test_tags) + ": Excludes the rest")\n')
                    f.write('\n\n')
        f.close()


def get_generate_data_exl(reportpath, sheetname, testcase, tendpoint, method, payload, apiresponse, schvariables,
                          sfvalidations, runtype, case, exresults, rpri, get_total_rows_x=0):
    reportpath.set_sheet_by_name(sheetname)
    get_total_rows = reportpath.get_max_row_no()
    try:
        if get_total_rows == get_total_rows_x:
            get_total_rows = 0
        reportpath.write_cell_content_colored(get_total_rows + 1, 1, testcase)
        reportpath.write_cell_content_colored(get_total_rows + 1, 2, tendpoint)
        reportpath.write_cell_content_colored(get_total_rows + 1, 3, method)
        reportpath.write_cell_content_colored(get_total_rows + 1, 4, payload)
        reportpath.write_cell_content_colored(get_total_rows + 1, 5, apiresponse)
        reportpath.write_cell_content_colored(get_total_rows + 1, 6, str(schvariables))
        reportpath.write_cell_content_colored(get_total_rows + 1, 7, str(sfvalidations))
        reportpath.write_cell_content_colored(get_total_rows + 1, 8, runtype)
        reportpath.write_cell_content_colored(get_total_rows + 1, 9, case)
        reportpath.write_cell_content_colored(get_total_rows + 1, 10, exresults)
        reportpath.write_cell_content_colored(get_total_rows + 1, 11, str(rpri))
    except Exception as e:
        if ValueError:
            pass


def get_sheetnames_excel(generate_test_ap, itemdata='COE_001', ustory=None):
    get_excel_data, gtdata = None, []
    getxcl = ParseExcel(generate_test_ap)
    getallxm = getxcl.get_sheetnames()
    if itemdata:
        getdata_item = itemdata.split(',')
        getdata_item.append(ustory)
        for gdata in range(0, len(getdata_item)):
            for xallm in getallxm:
                get_excel_data = ParseExcel(generate_test_ap).get_row_all_col_data(sheetname=xallm,
                                                                                   columnname="TC_No",
                                                                                   pagename=getdata_item[gdata])
                if get_excel_data:
                    gtdata.append(get_excel_data)
    else:
        for xallm in getallxm:
            get_excel_data = ParseExcel(generate_test_ap).get_row_all_col_data(sheetname=xallm, columnname="TC_No",
                                                                               pagename=ustory)
            if get_excel_data:
                gtdata.append(get_excel_data)
    return gtdata


def excel_row_data(gtdata, depend):
    testcase = gtdata["Test_Case"]
    testdef = gtdata["TC_No"]
    xendpoint = gtdata["EndPoint"]
    xmethod = gtdata["Method"]
    payload = gtdata["Payload"]
    params = gtdata["Params"]
    testdata = gtdata["Test_Data"]
    scvarib = gtdata["Assertion_key"]
    ctype = gtdata["Format_Validation"]
    exresults = gtdata["Expected_API_Response"]
    sprints = gtdata["IncidentId"]
    if depend is not None and testdef in depend:
        exestp = None
    else:
        exestp = 'Y'
    return testcase, xendpoint, xmethod, payload, params, testdata, scvarib, \
        ctype, exresults, sprints, exestp


def id_generator(size=6, chars=string.ascii_uppercase + string.digits):
    return ''.join(random.choice(chars) for _ in range(size))


def mixedcharacters(k=5):
    mixed_chars = ''.join(random.choices(string.ascii_letters + string.digits, k=k))
    return mixed_chars


def getlister(xnum, data=None, datex=None):
    my_list = []
    x = xnum
    for i in range(x):
        if data:
            char_type = data
        else:
            char_type = testdata.username() + id_generator(4) + '_' + str(datex)
        my_list.append(char_type)
    return my_list


def getExtracterx(data, xdata, tdata):
    getdata = data_required(datafile=data)
    if tdata == '!res' or '@' in tdata:
        datapack = Validations().get_json_string_results('data.' + str(xdata), getdata)
        if '.item.' in xdata:
            getxdata = datapack
        else:
            getxdata = str(getdata['data'][xdata])
    else:
        datapack = Validations().get_json_string_results(xdata, getdata)
        getxdata = datapack[0]
    return getxdata


def getExtract(xdata, tdata):
    fpath, aName, apiname, evalue = xdata.split('|')
    # print(fpath, aName, apiname, evalue)
    apiname = str(apiname) + '.json'
    getfile = mypath / 'test_data' / 'json_data' / fpath / apiname
    getdata = data_required(datafile=getfile)
    datapack = Validations().get_json_string_results(aName, getdata)
    if evalue == '':
        return datapack
    else:
        return datapack[0]


def getremove(xdata, tdata, getn=None):
    if '|' in xdata:
        fpath, aName, apiname, atname = xdata.split('|')
        apiname = str(apiname) + '.json'
        getfile = mypath / 'test_data' / 'json_data' / fpath / apiname
        getdata = data_required(datafile=getfile)
    else:
        getdata = getn
        aName = xdata
    datapack = Validations().get_json_string_results(aName, getdata)
    return datapack


def getbaseUrl(ddata, urlString, repl=None):
    match = re.search(r'{(.*?)}', ddata)
    if match:
        extracted_string = match.group(1)
        replaced_string = ddata.replace(str('{' + extracted_string + '}'), str(urlString))
        return replaced_string
    else:
        replaced_string = ddata.replace(str(repl), str(urlString))
        return replaced_string


def getepoint(getnbx):
    global progress, epointx
    config.read(mypath / "utils" / "endpointapi.ini")
    getslx = getnbx  # .split('/')
    for xlx in getslx[::-1]:
        timeout_start = time.time()
        timeout = 1 * 2
        progress = None
        while not progress:
            delta = time.time() - timeout_start
            try:
                progress = config.get("apiEndpoint", xlx)
                epointx = xlx
            except Exception:
                progress = None
            if progress is not None:
                progress = epointx
            if delta >= timeout:
                break
        if progress is not None:
            progress = epointx
            break
    return progress


def create_excel_sheet_results(file_path, hTitle, sHtitle):
    # Create a new workbook and select the active sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set the sheet name as "Object_repo"
    sheet.title = "API_Results"

    # Headers for the Excel sheet
    headers = ['Tc_no', 'API', 'Testcase', 'Method:Endpoint', 'Payload', 'Params',
               'Expected', 'Actual', 'Results', 'requestValidation', 'httpStatusCode', 'responseTime(Secs)',
               'RequirementId', 'IncidentId']

    # Set header values in the first row

    # Set formatting for the title row (first row)
    title_font = Font(name='Tahoma', size=20, bold=True, color="FFFFFF", underline="single")
    title_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    title_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    title_border = Border(top=Side(border_style="thin", color="000000"),
                          left=Side(border_style="thin", color="000000"),
                          right=Side(border_style="thin", color="000000"),
                          bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the title row (first row)
    sheet.merge_cells('A1:N1')

    # Apply formatting for the title row (first row)
    title_cell = sheet['A1']
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = title_alignment
    title_cell.border = title_border
    title_cell.value = hTitle

    # Set formatting for the subtitle row (second row)
    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")
    subtitle_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    subtitle_border = Border(top=Side(border_style="thin", color="000000"),
                             left=Side(border_style="thin", color="000000"),
                             right=Side(border_style="thin", color="000000"),
                             bottom=Side(border_style="thin", color="000000"))

    # Merge and center cells for the subtitle row (second row)
    sheet.merge_cells('E2:J2')

    # Apply formatting for the subtitle row (second row)
    subtitle_cell = sheet['E2']
    subtitle_cell.font = subtitle_font
    subtitle_cell.fill = subtitle_fill
    subtitle_cell.alignment = subtitle_alignment
    subtitle_cell.border = subtitle_border
    subtitle_cell.value = sHtitle

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    # Apply formatting for the subtitle row (second row)
    subtitle_cella = sheet['A2']
    subtitle_cella.font = subtitle_font
    subtitle_cella.fill = subtitle_fill
    subtitle_cella.alignment = subtitle_alignment
    subtitle_cella.border = subtitle_border
    subtitle_cella.value = 'Total Tests:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
    subtitle_cellb = sheet['B2']
    subtitle_cellb.font = subtitle_font
    subtitle_cellb.fill = subtitle_fill
    subtitle_cellb.alignment = subtitle_alignment
    subtitle_cellb.border = subtitle_border
    subtitle_cellb.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_cellc = sheet['C2']
    subtitle_cellc.font = subtitle_font
    subtitle_cellc.fill = subtitle_fill
    subtitle_cellc.alignment = subtitle_alignment
    subtitle_cellc.border = subtitle_border
    subtitle_cellc.value = 'Total Passed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
    subtitle_celld = sheet['D2']
    subtitle_celld.font = subtitle_font
    subtitle_celld.fill = subtitle_fill
    subtitle_celld.alignment = subtitle_alignment
    subtitle_celld.border = subtitle_border
    subtitle_celld.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_cellh = sheet['K2']
    subtitle_cellh.font = subtitle_font
    subtitle_cellh.fill = subtitle_fill
    subtitle_cellh.alignment = subtitle_alignment
    subtitle_cellh.border = subtitle_border
    subtitle_cellh.value = 'Total Failed:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="FC2C03", end_color="FC2C03", fill_type="solid")
    subtitle_celli = sheet['L2']
    subtitle_celli.font = subtitle_font
    subtitle_celli.fill = subtitle_fill
    subtitle_celli.alignment = subtitle_alignment
    subtitle_celli.border = subtitle_border
    subtitle_celli.value = '0'

    subtitle_font = Font(name='Tahoma', size=12, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellj = sheet['M2']
    subtitle_cellj.font = subtitle_font
    subtitle_cellj.fill = subtitle_fill
    subtitle_cellj.alignment = subtitle_alignment
    subtitle_cellj.border = subtitle_border
    subtitle_cellj.value = 'Total Skipped:'

    subtitle_font = Font(name='Tahoma', size=24, bold=True, color="FFFFFF", underline="single")
    subtitle_fill = PatternFill(start_color="F4D25A", end_color="F4D25A", fill_type="solid")
    subtitle_cellk = sheet['N2']
    subtitle_cellk.font = subtitle_font
    subtitle_cellk.fill = subtitle_fill
    subtitle_cellk.alignment = subtitle_alignment
    subtitle_cellk.border = subtitle_border
    subtitle_cellk.value = '0'

    # Set formatting for the headers (third row)
    header_font = Font(bold=True, color="FFFFFF", underline="single")
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid")

    sheet.append(headers)
    # Apply formatting for the headers (third row)
    for cell in sheet['A3:N3'][0]:
        cell.font = header_font
        cell.fill = header_fill

    # Apply borders to the first two rows except the header row
    row_border = Border(top=Side(border_style="thin", color="000000"),
                        left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"))

    for row in sheet.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.border = row_border

    # Save the Excel file
    workbook.save(file_path)