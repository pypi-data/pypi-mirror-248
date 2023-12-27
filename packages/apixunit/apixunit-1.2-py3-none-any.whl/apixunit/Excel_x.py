from apixunit.CONST_j import CONST
from pathlib import Path
import openpyxl
import shutil
import random

import time
mypath = Path.cwd()

class Excel():
    def __init__(self, path=CONST.EXCELPATH):
        try:
            self.excelwb = openpyxl.load_workbook(path)
            self.excelst = ''
            self.filepath = path
        except Exception as msg:
            print(msg)

    def Select_Sheet_By_Name(self, name):
        if name in self.excelwb.sheetnames:
            self.excelst = self.excelwb[name]
            return self.excelst
        else:
            raise Exception("Cannot find the given sheet name - %s" % name)

    def Get_Row_Numbers(self):

        rows = self.excelst.max_row

        for i in range (1, rows + 1):
            if self.excelst.cell(row=i,column=1).value is not None:
                #print(self.excelst.cell(row=i,column=1).value)
                if i != rows:
                    continue
                else:
                    return i
            else:
                return i - 1

    def Get_All_Values_By_ColName(self, colname):

        values = []
        for c in range(1, self.excelst.max_column + 1):
            if str(self.excelst.cell(row=1, column=c).value).lower() == colname.lower():
                for r in range(2, self.Get_Row_Numbers() + 1):
                    if self.excelst.cell(row=r, column=c).value is not None:
                        values.append(self.excelst.cell(row=r, column=c).value)
        if len(values) != 0:
            return values
        else:
            raise Exception("Cannot get any value by using the given column name - %s" % colname)

    def Get_Value_By_ColName(self, colname, row, path = ''):

        for c in range(1, self.excelst.max_column + 1):
            if str(self.excelst.cell(row=1, column=c).value).lower() == colname.lower():
                if row <= self.excelst.max_row:
                    value = self.excelst.cell(row=row + 1, column=c).value

                    if str(value).find("random") != -1 and path != "" and str(value).lower() != "randomindex":
                        excelTemp = Excel(CONST.EXCELPATH)
                        excelTemp.Select_Sheet_By_Name("data")
                        randomlist = excelTemp.Get_All_Values_By_ColName(value)
                        randomvalue = random.choice(randomlist)
                        file = Path(path) / Path("%d_%s_%s.random" % (row, colname, str(randomvalue)))
                        file.touch()
                        #print(str(file))
                        #self.excelst.cell(row=row + 1, column=c).value = "%s(%s)" % (str(value), str(randomvalue))
                        return randomvalue
                    else:
                        return value
                else:
                    raise Exception("Given row number %d is larger than the rows of the sheet" % row)
        else:
            raise Exception("Cannot find the given column name - %s" % colname)

    def Set_Value_By_ColName(self, content, colname, row, sheetpassed=''):
        if sheetpassed != '':
            sheet = sheetpassed
        else:
            sheet = self.excelst

            for c in range(1, sheet.max_column + 1):
                if str(sheet.cell(row=1, column=c).value).lower() == colname.lower():
                    if row <= sheet.max_row:
                        sheet.cell(row=row + 1, column=c).value = content
                        break
                    else:
                        raise Exception("Given row number %d is larger than the rows of the sheet" % row)
            else:
                raise Exception("Cannot find the column name - %s" % colname)

    def Set_Sheet_Name(self,old_name,new_name):
        if old_name in self.excelwb.sheetnames:
            self.excelwb[old_name].title = new_name
        else:
            raise Exception("Cannot find the given sheet name - %s" % old_name)

    def Get_Excution_DataSet(self,colname,dataset_name="Data set",keyword="Not Yet"):
        DaList = []
        for c in range(1, self.excelst.max_column + 1):
            if str(self.excelst.cell(row=1, column=c).value).lower() == colname.lower():
                for r in range(1,self.excelst.max_row + 1):
                    if str(self.excelst.cell(row=r + 1, column=c).value).lower() == keyword.lower():
                        if self.Get_Value_By_ColName(dataset_name,r) != '':
                            #print("The value of cell(%d,%d) is %r" % (r,c, self.Get_Value_By_ColName(dataset_name,r)))
                            DaList.append(int(self.Get_Value_By_ColName(dataset_name,r)))
        if len(DaList) != 0:
            return DaList
        else:
            raise Exception("Not found any executable %s by using the %s keyword" % (dataset_name, keyword))

    def Save_Excel(self,path=''):
        if path == '':
            savepath = self.filepath
        else:
            savepath = path

        try:
            self.excelwb.save(savepath)
        except Exception as msg:
            print(msg)

