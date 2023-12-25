"""
TCMS Import/Export Function
Usage: 
    kiwi_tcms.py login --tcms_host=<IP> --username=<username> --password=<password>
    kiwi_tcms.py import [--path=<path_to_xls_file>]
    kiwi_tcms.py export [--path=<path_to_xls_file>]
    kiwi_tcms.py result --run_id=<id>
    kiwi_tcms.py remove --object=<component,product> --value=<value> 


Arguments:
  Options:
  -h --help
"""

import sys
import os, docopt, getpass
import string
from unicodedata import name
import urllib.parse as urlparse
import glob
import json, csv
from tcms_api import TCMS
from openpyxl import load_workbook

#sys.setdefaultencoding('utf-8')

#ssl.SSLContext.verify_mode = ssl.VerifyMode.CERT_OPTIONAL

#tc_priority = {"P0": 6, "P1": 1, "P2": 2, "P3": 3}
#tc_priority = {"P0": 29, "P1": 10, "P2": 15, "P3": 23}

product_name = 'Flexiwan' #{"Flexiwan": 1}
#product = 5388 #{"Flexiwan": 1}

#categories = {"Functionality": 2, "Interoperability": 3, "Stress": 4, "Scale": 5, "Performance": 6, "Upgrade": 7, "Downgrade": 8}
#categories = {"Functionality": 5821, "Interoperability": 5903, "Stress": 5904, "Scale": 5905, "Performance": 5906, "Upgrade": 5907, "Downgrade": 5908}

#components = {"flexiAgent": 1, "flexiManage": 2, "flexiRouter": 3, "flexiEdge": 4}
#components = {"flexiAgent": 154, "flexiManage": 155, "flexiRouter": 156, "flexiEdge": 157}

#ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = "/etc"
XLS_PATH = glob.glob(os.path.join(ROOT_DIR, 'xls/')+"*.xlsx")
CSV_PATH = os.path.join(ROOT_DIR, 'csv/')

class kiwi_tcms():
    #rpc_client = class()
    def __init__(self):
        self.rpc_client = TCMS()

    def config_fname(self):
        return os.path.expanduser(ROOT_DIR + "/" + "tcms.conf")

    def tcms_url(self, tcms, path):
        return urlparse.urlunsplit(("https", tcms, path, "", "" ))

    def login(self, args):
        tcms_host = args["--tcms_host"]
        url = self.tcms_url(tcms_host, "/xml-rpc/")
        user = args["--username"]
        passwd = args["--password"]
        if not user: user = input("Login: ")
        if not passwd: passwd = getpass.getpass("Password: ")
        with os.fdopen(os.open(self.config_fname(), os.O_WRONLY | os.O_CREAT, 600), 'w') as conf :
            conf.write(f"[tcms]\nurl = {url}\nusername = {user}\npassword = {passwd}")
        self.rpc_client = TCMS()


    def readXls(self, pathOfXls, sheetName):
        output = []
        wb = load_workbook(pathOfXls)
        sheet = wb[sheetName]
        for row in sheet.iter_rows(2, sheet.max_row):
            list = []
            if row[0].value is not None:
                for col in range(sheet.max_column):
                    list.append(row[col].value)
                #print(list)
                output.append(list)
        return output

    def importXls(self, args):
        if not args["--path"]:
            PATH=XLS_PATH
        else:
            PATH=glob.glob(os.path.join(args["--path"])+"*.xlsx")
        tot=0
        testCases = []
        print(PATH)
        for file in PATH:
            #readXls1 = ReadXls()
            print(f"File name to be Imported: {file}")
            testCases = self.readXls(file, "Test Cases")
            tc = self.createTestCases(testCases)
            tot = tot + len(testCases)
        print(f"{tot} Test Cases were Imported Successfully")
    
    def exportCsv(self, args):
        if not args["--path"]:
            path=CSV_PATH
        else:
            path = args["--path"]
        tag_id = self.rpc_client.exec.Tag.filter({'name__in': ['IKEv2']})[0].get("id")
        tc_list = self.rpc_client.exec.TestCase.filter({'tag__in': [tag_id]})
        csv_file = open(path + 'export.csv', 'w')
        csv_object = csv.writer(csv_file)
        count = 0
        for tc in tc_list:
            if count == 0:
                header = tc.keys()
                csv_object.writerow(header)
                count =+ 1
            csv_object.writerow(tc.values())
        csv_file.close()

    
    # def validateSheets(self):
    #     testcases = self.xlsToList()
    #     category_idx = priority_idx = summary_idx = 0
    #     precondition_idx = expected_results_idx = steps_idx = 0
    #     for idx, columns in enumerate(testcases[0]):
    #         print(idx, columns)
    #         if columns[idx] == 'Categories':
    #             category_idx = idx
    #         elif columns[idx] == 'Summary':
    #             summary_idx = idx
    #         elif columns[idx] == 'Precondition':
    #             precondition_idx = idx

    def createTestCases(self, test_cases):
        #print(self.rpc_client.exec.Product.filter({'name': product_name})[0].get("id"))
        product_id = self.rpc_client.exec.Product.filter({'name': product_name})[0].get("id")
        print(f"Product ID is : {product_id}")
        for id, case in enumerate(test_cases):
            #category = categories.get(case[0])
            category_name = case[0]
            category = self.rpc_client.exec.Category.filter({'name': case[0]})[0].get("id")
            #category = case[0]
            feature = case[1]
            summary = case[2]
            precondition = case[3]
            steps = case[4]
            expected_results = case[5]
            is_automated = 'True' if case[9] == 'True' else 'False'
            #priority = tc_priority.get(case[6])
            priority_name = case[6]
            priority = self.rpc_client.exec.Priority.filter({'value': case[6]})[0].get("id")
            #priority = case[6]
            component = case[7]
            text = f'**Precondition: ** \n{precondition} \n\n\n**Steps: ** \n{steps} \n\n\n **Expected :** \n{expected_results}'
            test_case = {
                'summary': summary,
                'text': text,
                'product': product_id,
                'category': category,
                'priority': priority,
                'case_status': 1,  # PROPOSED
                'is_automated': is_automated,
                'managers_of_runs': 'on',
                'assignees_of_case_runs': 'on',
                'default_tester_of_case': 'on',
                'default_testers_of_runs': 'on',
                'notify_on_case_update': 'on',
                'notify_on_case_delete': 'on'
            }
            print(test_case)
            tc = self.rpc_client.exec.TestCase.create(test_case)
            #print(tc)
            print(f"test case is added and it is {tc}\n")
            self.rpc_client.exec.TestCase.add_component(tc['id'], component)
            self.rpc_client.exec.TestCase.add_tag(tc['id'], feature)
            #return tc

    def removeTestCases(self, args):
        object = args["--object"]
        value = args["--value"]
        if object == 'tag':
            object_id = self.rpc_client.exec.Tag.filter({'name__in:'[value]})[0].get("id")
        elif object == 'component':
            object_id = self.rpc_client.exec.Component.filter({'name__in:'[value]})[0].get("id")
        elif object == 'category':
            object_id = self.rpc_client.exec.Category.filter({'name__in:'[value]})[0].get("id")

        print(object_id)
        #query = {'tag__in': ['496']}
        query = {f'{object}__in':[object_id]}
        tc_list = self.rpc_client.exec.TestCase.filter(query)
        tc_remove = self.rpc_client.exec.TestCase.remove(query)
        print(f"{len(tc_list)} Test Case are removed")

        #print(tc_remove)

    def exportResult(self, args):
        tr = args["--run_id"]
        tc_list = list()
        tr_file = "status.csv"
        tc_list = self.rpc_client.exec.TestRun.get_cases(int(tr))
        #print(tc_list[1])
        with os.fdopen(os.open(tr_file, os.O_WRONLY | os.O_CREAT, 600), 'w', encoding="utf-8") as conf :
            conf.write(f"ID, Feature, Summary, Automated, Author, Priority, TR-{tr} Status\n")
            for tc in tc_list:
                #tcs = f"{tc['id']}, {tc['is_automated']} \n"
                #print(str(tc) + "\n")
                user_id = tc['default_tester']
                tc_id = tc['id']
                user_query = {f'id__in':[user_id]}
                tc_execution = {f'case__in':[tc_id]}
                tag = self.rpc_client.exec.Tag.filter(tc_execution)[0].get("name")
                summary = str(tc['summary']).translate(str.maketrans('', '', string.punctuation))
                ##summary = ''.join(summary.splitlines())
                #print(tag)
                username = self.rpc_client.exec.User.filter(user_query)[0].get("username")
                #print(username)
                #print(f"{tc['id']}, {str(tag)}, {summary}, {tc['is_automated']}, {username}, {tc['priority']}, {tc['status']} \n")
                conf.write(f"{tc['id']}, {tag}, {summary}, {tc['is_automated']}, {username}, {tc['priority']}, {tc['status']} \n")


t = kiwi_tcms()
_function_map = {
    "login": t.login,
    "import": t.importXls,
    "export": t.exportCsv,
    "result": t.exportResult,
    "remove": t.removeTestCases
}

if __name__ == "__main__":
    args = docopt.docopt(__doc__)
    try:
        _function_map[sys.argv[1]](args)
    except RuntimeError as e:
        print(sys.stderr, e.message)
        sys.exit(1)